from openpyxl import Workbook

wb = Workbook()

# Sheet 1
ws1 = wb.active
ws1.title = "Financial Metrics"

with open("financial_metrics.txt", "r", encoding="utf-8") as f:
    data = f.read()

ws1["A1"] = data

# Sheet 2
ws2 = wb.create_sheet("Comparison")

with open("comparison_report.txt", "r", encoding="utf-8") as f:
    data = f.read()

ws2["A1"] = data

# Sheet 3
ws3 = wb.create_sheet("Recommendation")

with open("final_recommendation.txt", "r", encoding="utf-8") as f:
    data = f.read()

ws3["A1"] = data

wb.save("research_report.xlsx")

print("Excel report generated.")